#!/usr/bin/python

# region import
import os
import sys  # potrzebna do sprawdzenia, czy użytkownik podał jakiekolwiek parametry na wejściu
import argparse  # do obsługi parametrów wejściowych
import statistics

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

import geocoder
from ui_functions import *
from weatherApis import *
# endregion

# region zmienne startowe
CITY = "gdynia"  # nazwy miast z małych liter aby łatwiej było operać na API
# współrzedne dla Gdyni pobrane z portalu https://www.wspolrzedne-gps.pl/
LON = 54.5142351  # długość geograficzna
LAT = 18.5358849  # szerokość geograficzna

MODE = "current"  # "current" podaję aktualne dane, alternatywny tryb -> "forecast" - prognoza, ale tylko dla pogody

LANG = "pl"  # język do komunikacji z API
DAYS = 5  # ilość dni do przodu na które można uzyskać prognoze pogody
SILENT = False

weatherDataset = []

option = None  # zmienna określająca, którą opcję wybrał uzytkownik
is_start_with_parameters = False  # flaga informująca czy program został uruchomiony z parametrami
error_in_start_parameters = 0  # zmienna określająca czy wszystkie parametry zostały podane poprawnie przez Użytkownika
# endregion

# wyświetl wiadomość powitalną
welcome_message()

# sprawdź ilość argumentów podanych na starcie przez użytkownika
#  zawsze jest minimum 1 - nazwa skryptu. Jeżeli ejst ich więcej to znaczy, ze użytkownik podał parametry startowe.
#  Jeżeli tylko 1 to znaczy, że uruchomił skrypt bez parametrów
number_of_arguments = len(sys.argv)

# jeżeli użytkownik podał jakiekolwiek argumenty to rozpocznij działanie programu na parametrach

# region parametry startowe przechwytywanie i wstępna obsługa
if number_of_arguments > 1:

    is_start_with_parameters = True

    # region przechwytywanie parametrów
    parser = argparse.ArgumentParser()

    # deklaracja parametru startowego --city (nazwy miasta). nargs z wartością "*" pozwala na pobranie 1 lub więcej
    # wartości dla danego parametru - tutaj potrzebne, bo nazwy miast mogą być wielowyrazowe
    # wartość domyślna zapisana w zmiennych globalnych na początku programu zamiast tutaj w parserze, aby uniknąć
    #  późniejszych problemów
    parser.add_argument("-c", "--city", nargs="*", dest="city", default=None,
                        help="pobiera nazwę miasta. Wartość domyślna: Gdynia")

    # deklaracja parametru startowego --longitude (długości geograficznej)
    # wartość domyślna (54.23) zapisana w zmiennych globalnych na początku programu zamiast tutaj w parserze, aby
    #  uniknąć późniejszych problemów
    parser.add_argument("-lon", "--longitude", nargs="*", dest="longitude", default=None,
                        help="pobiera wartość długości geograficznej w formie liczby całkowitej lub ułamka "
                             "dziesiętnego. Pamiętaj, że część całości od ułamka oddziela '.' (KROPKA) ! "
                             "Wartość domyślna: 54.23 (dla Gdyni).")

    # deklaracja parametru startowego --latitude (szerokości geograficznej)
    # wartość domyślna zapisana w zmiennych globalnych to 19.23
    parser.add_argument("-lat", "--latitude", nargs="*", dest="latitude", default=None,
                        help="pobiera wartość szerokości geograficznej w formie liczby całkowitej lub ułamka "
                             "dziesiętnego. Pamiętaj, że część całości od ułamka oddziela '.' (KROPKA) ! "
                             "Wartość domyślna: 19.23 (dla Gdyni).")

    # parametr decydujące, czy program ma działać w trybie "current", czy "forecast"
    parser.add_argument("-m", "--mode", dest="mode", choices=['current', 'forecast'], default=None,
                        help="ustawia tryb działania programu - 'current' pokazuje aktualne dane,"
                             "'forecast' pokazuje prognozę (tylko dla pogody!)")

    # parametr informujący na ile dni do przodu ma pokazać prognozę pogody (zakres 1-6, bo 6 się nie wlicza)
    parser.add_argument("-d", "--days", dest="days", type=int, choices=range(1, 6), default=5,
                        help="ustala na ile dni do przeodu ma wyświetlić progrnozę pogody. Działa tylko z"
                             "'--mode=forecast'. Trybie danych bieżących (--mode=current) parametr będzie ignorowany.")

    # tworzy słownik argumentów
    args = parser.parse_args()
    # endregion

# region gdy użytkownik spróbował jednocześnie podać nazwę miasta oraz namiary GPS w parametrach startowych...
    if (args.city is not None) and ((args.longitude is not None) or (args.latitude is not None)):

        # wyłączam opcję startu z parametrami
        is_start_with_parameters = False

        print("""
OSTRZEZENIE! Podjęto próbę jednoczesnego podania: nazwy miasta i współrzędnych GPS!

Co chcesz zrobić?
    1. Skorzystać z podanej nazwy miejscowości.
    2. Skorzystać z podanych współrzędnych GPS.
    3. Przejść do menu startowego.
    4. Zakończyć działanie programu.
""")

        # wykonuj pętlę dopóki użytkownik nie poda prawidłowego numeru opcji
        while option is None:

            # użytkownik wporwadza numer komendy - numer jest w postaci string, aby uniknąć błędów związanych z podaniem
            #  nieprawdiłowego znaku lub ciągu znaków
            option = check_user_choice_is_correct(input("Wprowadź numer opcji (1/2/3/4) i naciśnij ENTER: "),
                                                  ["1", "2", "3", "4"])
            print("")  # nowa linia dla poprawy czytelności

            # jeżeli nie została zwrócona żadna opcja to znaczy, że użytkownik podał, opcję spoza dopuszczonych
            if option is None:
                print("Nie ma takiej opcji! Spróbuj jeszcze raz...\n")

        # reaguj na prawidłowo wybraną opcję
        if option == "1":
            # wyzeruj zmienne -lon i -lat
            args.longitude = None
            args.latitude = None

        elif option == "2":
            # wyczyść nazwę miasta
            args.city = None

        elif option == "3":
            # ustawienie number_of_arguments na 0 spowoduje wywołanie menu start w dalszej części programu poprzez
            #  zasymulowanie jakby argumenty nie zostały podane (powinno być 1, ale podanie wartości 0 zasugeruje, że
            #  to jest sztucznie wytworzona zmiana)
            number_of_arguments = 0

            # wyczyść wszystkie dane podane w parametrach przez użytkownika
            args.city = None
            args.longitude = None
            args.latitude = None

        elif option == "4":
            print("Zakończono działanie programu.")
            exit()

        else:
            print("Coś poszło nie tak!")
            exit()  # kończy natychmiast program aby uniknąć potencjalnych błedów z powodu wybrania opcji, której nie ma
# endregion

# region został wywołany parametr -city
    #  to przepisz jego wartość do zmiennej globalnej CITY, a jak nie została podana nowa wartość dla niego (domyślnie
    #  w parserze argumentów startowych jest None) to pomiń ten krok i pozostaw niezmienioną wartość CITY (zgodną
    #  z wartością domyslną podaną dla zmiennych globalnych na poczatku)
    if (args.city is not None) and (args.longitude is None) and (args.latitude is None):
        # wyczyść nazwę miasta
        CITY = ""

        # przejdź przez całą listę wyrazów, z których może się składać nazwa miasta. Miasto mogą być 1 wyrazowe
        #  np.: Gdynia, Sopot, Wejherowo, albo wielowyrazowe: Stalowa Wola, Kędzierzyn Koźle
        for word_of_city_name in args.city:

            # dodawaj kolejne wyrazy do zmiennej CITY (zmienna typu string)
            CITY = CITY + word_of_city_name + " " 

        # oczyszczanie zmiennej z białych znaków na początku i na końcu
        CITY = CITY.strip()

        LON, LAT = get_coords_by_city_name(CITY, "", is_start_with_parameters)

        error_in_start_parameters = 0
# endregion

# region jeżeli zostały wywołane parametry -lon i -lat
    #  to pobierz dane z tych parametrów i przekaż je do zmiennych globlanych LON i LAT. Podanie przez użytkownika
    #  większej ilości liczb dla danego parametru (np.: -lon 34.5 78 43.2) będzie zignorowane i zostanie pobrana tylko
    #  pierwsza wartość (tu w przykładzie 34.5)
    elif (args.city is None) and (args.longitude is not None) and (args.latitude is not None):

        # sprawdzam, czy wartości podane przez użytkownika z pomocą parametrów są prawidłowe
        is_correct_value_LON, is_correct_range_LON, LON = check_GPS_value_from_user(args.longitude[0], 0, 180)
        is_correct_value_LAT, is_correct_range_LAT, LAT = check_GPS_value_from_user(args.latitude[0], 0, 90)

        # jeżeli wszystko jest ok to ustawiam flagę error_in_start_parameters na 0
        if (is_correct_value_LON == 1 and
                is_correct_range_LON == 1 and
                is_correct_value_LAT == 1 and
                is_correct_range_LAT == 1):

            error_in_start_parameters = 0

            # pobierz miejsce na podstawie koordynatów GPS
            g = geocoder.geocoder()
            place = g.getQueryReverseResults(LAT, LON)

            # jeżeli nie nie zwróciło żadnego miasta
            if place['city'] == '':
                CITY = 'Brak miasta w tym miejscu'
            else:
                CITY = place['city']

        # jeżeli jest jakikolwiek błąd przy sprawdzaniu poprawności to ustawiam error_in_start_parameters na 1
        else:
            error_in_start_parameters = 1

    else:
        error_in_start_parameters = 1
# endregion

# region jeżeli został wywołany parametr --mode
    if args.mode is not None:
        MODE = args.mode
        if args.mode == 'forecast':
            DAYS = args.days
# endregion

# region obsługa błędnie podanych parametrów

    # wyświetl info o błędzie we współrzędnych wtedy gdy taki błąd znajdziesz oraz gdy użytkownik rzeczywiście
    #  podał jakikolwiek argument
    if error_in_start_parameters == 1 and number_of_arguments > 1:
        print("""
BŁĄD! Podana tylko jedna współrzędna GPS lub współrzędne podane niepoprawnie (wartość nieliczbowa bądź poza zakresem).

Co chcesz zrobić?
    1. Przejść do menu startowego.
    2. Zakończyć działanie programu.
""")

        # zerowanie
        option = None

        # wykonuj pętlę dopóki użytkownik nie poda prawidłowego numeru opcji
        while option is None:
            option = check_user_choice_is_correct(input("Wprowadź numer opcji (1/2) i naciśnij ENTER: "), ["1", "2"])
            print("")  # nowa linia dla poprawy czytelności

            # sprawdź, czy podana opcja jest poprawna, jeżeli nie to wyświetl komunikat o złym wyborze
            if option is None:
                print("Nie ma takiej opcji! Spróbuj jeszcze raz...\n")

        # reaguj na prawidłowo wybraną opcję
        if option == "1":
            # ustawienie number_of_arguments na 0 spowoduje wywołanie menu start w dalszej części programu poprzez
            #  zasymulowanie jakby argumenty nie zostały podane (powinno być 1, ale podanie wartości 0 zasugeruje, że
            #  to jest sztucznie wytworzona zmiana)
            number_of_arguments = 0

            # wyczyść wszystkie dane podane w parametrach przez użytkownika
            args.city = None
            args.longitude = None
            args.latitude = None

        elif option == "2":
            print("Zakończono działanie programu")
            exit()

        else:
            print("Coś poszło nie tak!")
            exit()  # kończy natychmiast program aby uniknąć potencjalnych błedów z powodu wybrania opcji, której nie ma
# endregion

# region menu startowe
# Jeżeli użytkownik nie podał argumentów, albo podał je błędnie number_of_arguments == 0) to rozpocznij program od
#  menu startowego
if number_of_arguments <= 1:

    # menu startowe
    print("""    
Co chcesz zrobić?
  1. Pokaz aktualne dane.
  2. Pokaż prognozę pogody.
  3. Zakończ działanie programu.
""")

    option = None  # resetuję zmienną

    # wykonuj pętlę dopóki użytkownik nie poda prawidłowego numeru opcji
    while option is None:
        option = check_user_choice_is_correct(input("Wprowadź numer opcji (1/2/3) i naciśnij ENTER: "),
                                              ["1", "2", "3"])
        print("")  # nowa linia dla poprawy czytelności

        # sprawdź czy podana opcja jest poprawna, jeżeli nie to wyświetl komunikat o złym wyborze
        if option is None:
            print("Nie ma takiej opcji! Spróbuj jeszcze raz...\n")

    # dalej zbieram potrzebne informacje do wyświetlenia danych
    if option == "1":
        MODE = "current"  # tryb pokazywania aktualnych danych
        print("Aktualne dane")

        # funkcja pobiera zmienne globalne, użytkownik wprowadza miejsce, dla którego chce otrzymać informacje pogodowe
        #  i funkcja zwraca wskazane miejsce ponownie do zmiennych globalnych
        CITY, LON, LAT = get_place_from_user(CITY, LON, LAT)

    elif option == "2":
        MODE = "forecast"  # tryb pokazywania prognozy pogody
        print("Prognoza pogody")

        CITY, LON, LAT = get_place_from_user(CITY, LON, LAT)

    elif option == "3":
        print("Zakończono działanie programu.")
        exit()

    else:
        print("Coś poszło nie tak przy wybieraniu opcji!")
        exit()  # kończy natychmiast program aby uniknąć potencjalnych błedów z powodu wybrania opcji, której nie ma
# endregion

# jeżeli został wybrany tryb danych aktualnych to zmień ilość dni dla API na 1
if MODE == 'current':
    DAYS = 1

elif MODE == 'forecast':

    # zeruję wartość aby Użytkownik mógł sobie wybrać ilość dni
    DAYS = 0

    # dla poprawy czytelności
    print('')

    # czekaj na podanie wartości liczbowej w prawidłowym zakresie (od 1 do 5)
    while DAYS < 1 or DAYS > 5:
        try:
            DAYS = int(input("Na ile dni do przodu chcesz uzyskać prognozę pogody? (Zakres od 1 do 5): "))

        except ValueError:
            print('Podana wartość nie jest liczbą dziesiętną! Spróbuj jeszcze raz.\n')

        if DAYS < 1 or DAYS > 5:
            print("Nieprawidłowa wartość! Spróbuj jeszcze raz.\n")

else:
    print('Coś poszło nie tak...\nZakończono działanie programu.')
    exit()

# region kontrolny wydruk zmiennych podanych przez użytkownika
print("")
print("Pobieram dane dla następujących parametrów: ")
print("Miasto: ", CITY)
print("Długość geograficzna: ", LON)
print("Szerokość geograficzna: ", LAT)
print("Tryb pobierania danych: ", MODE)
if MODE == 'forecast':
    print("Z ilu następnych dni zostanie pobrana prognoza: ", DAYS)
print("")

# endregion

# region pobieranie danych z API

# check API keys to determine if the weather can be obtained(includning timezones)
if( check_API_keys(verify_online=False)==0 ):
  sys.exit("Nie skonfigurowano żadnego systemu do pobierania danych o pogodzie.\n Program nie może działać.\n Wpisz xxx -help, aby dowiedzieć się, jak dokonać konfiguracji.")


for API in os.listdir("config/API/"):
#for API in APIS:
  if API.endswith(".csv") and not API.startswith("."):
    #print(API)
    wa = WeatherApis()
    if wa.read_conf(API) == False:
      continue
    if wa.get_weather(MODE, LANG, DAYS, SILENT, LAT, LON):
      weatherDataset.append( wa.parse_result(MODE, DAYS) )
    else:
      print("Problem z uzyskaniem danych z "+wa.config["api_name"]+". Adres: "+wa.url_search)


def is_number(str):
    try:
        float(str)
        return True
    except:
        return False
################################################################## SAVE DATA TO EXCEL #############################################################

wb = Workbook()
dest_filename = 'results.xlsx'
ws = wb.active

CSV="api_name,"
row_xls0 = []
row_xls_empty = []
row_xls0.append("api_name")
for factor in factors:
    if(MODE=="forecast" and len(factorsDict[factor].type)==2 ) or (MODE=="current" and len(factorsDict[factor].type)==1):
        CSV = CSV +factor+","
        row_xls0.append(factor+" ["+factorsDict[factor].unit+"]" )
        row_xls_empty.append("")
CSV=CSV+"\r\n"
ws.append(row_xls0)

rowNum = 0

for day in range(0,DAYS):
    STAT = {}
    for factor in factors:
        STAT[factor]=[]
    for wynikiAPI in weatherDataset:
        rowNum  = rowNum + 1
        row_xls = []
        row_xls.append(str(wynikiAPI[day]["api_name"]))
        CSV = CSV + "\"" + str(wynikiAPI[day]["api_name"]) + "\","
        for factor in wynikiAPI[day]:
            if(MODE=="forecast" and len(factorsDict[factor].type)==2 ) or (MODE=="current" and len(factorsDict[factor].type)==1):
                CSV = CSV + "\"" + str(wynikiAPI[day][factor]) + "\","
                row_xls.append( str(wynikiAPI[day][factor]) )
                if( is_number(wynikiAPI[day][factor]) ):
                    STAT[factor].append( float(wynikiAPI[day][factor]) )
        CSV=CSV+"\r\n"
        ws.append(row_xls)
    CSV=CSV+"MEAN ± SD:,"


    #calculating statistics
    rowNum = rowNum + 2
    colNum = 1
    ws.cell(row=rowNum, column=colNum, value="MEAN ± SD:").fill = PatternFill(fgColor="FFFFFF", fill_type = 'solid')
    colNum = 2
    for factor in factors:
        if(MODE=="forecast" and len(factorsDict[factor].type)==2 ) or (MODE=="current" and len(factorsDict[factor].type)==1):
            if len(STAT[factor])>0:
                value = round(statistics.mean(STAT[factor]),1)
                stdev = round(statistics.pstdev(STAT[factor]),1)
                color = factorsDict[factor].convert_to_rgb(value)
                fill = PatternFill("solid", fgColor=color)
                ws.cell(row=rowNum, column=colNum, value=value).fill = PatternFill(fgColor=color, fill_type = 'solid')
                CSV=CSV+"\""+str(value)+"±"+str(stdev)+"\","
            else:
                CSV=CSV+"\"\","
            colNum = colNum + 1

    CSV=CSV+"\r\n\r\n"
    row_xls0[0]=""
    ws.append(row_xls_empty)
    ws.append(row_xls_empty)
    if(day==DAYS):
        ws.append(row_xls0)



#adjust columns
for col in ws.columns:
     max_length = 0
     column = col[0].column # Get the column name
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 1) * 1.05
     ws.column_dimensions[column].width = adjusted_width   

wb.save(filename = dest_filename)


f = open( 'results.csv', 'w' )
f.write( CSV )
f.close()

debug=0
if debug>0:
    with open('data.json', 'w') as outfile:
        for WD in weatherDataset:
            json.dump(WD, outfile)
            #print(json.dumps(WD, sort_keys=False, indent=4))